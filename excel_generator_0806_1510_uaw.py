# 代码生成时间: 2025-08-06 15:10:35
import os
from celery import Celery
from celery.utils.log import get_task_logger
from openpyxl import Workbook
from openpyxl.utils.exceptions import InvalidFileException
from openpyxl import load_workbook
from datetime import datetime

# Configure Celery
app = Celery('excel_generator', broker='pyamqp://guest@localhost//')
logger = get_task_logger(__name__)

@app.task
def generate_excel(data):
    """Generate an Excel file with the provided data.

    Args:
        data (dict): A dictionary containing the data to be written to the Excel file.
            Expected structure: {'sheet_name': ['row_data1', 'row_data2', ...]}
    
    Returns:
        dict: A dictionary containing the status of the Excel generation.
    """
    try:
        # Create a new workbook and add a worksheet
        wb = Workbook()
        ws = wb.active
        ws.title = 'Generated Data'
        datetime_str = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f'data_{datetime_str}.xlsx'
        
        # Write data into the sheet
        for row_index, row_data in enumerate(data.values(), start=1):
            for col_index, value in enumerate(row_data, start=1):
                ws.cell(row=row_index, column=col_index, value=value)

        # Save the workbook to a file
        wb.save(filename=filename)
        logger.info(f'Excel file {filename} generated successfully.')
        return {'status': 'success', 'filename': filename}
    
    except InvalidFileException as e:
        logger.error(f'Invalid Excel file: {e}')
        return {'status': 'error', 'message': str(e)}
    
    except Exception as e:
        logger.error(f'An error occurred: {e}')
        return {'status': 'error', 'message': str(e)}

# Example usage
if __name__ == '__main__':
    data = {
        'Sheet1': [
            ['Header1', 'Header2', 'Header3'],
            ['Data1', 'Data2', 'Data3'],
            ['Data4', 'Data5', 'Data6']
        ]
    }
    result = generate_excel.delay(data)
    print(result.get())